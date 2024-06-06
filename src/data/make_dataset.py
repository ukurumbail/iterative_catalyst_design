# -*- coding: utf-8 -*-
import os
import click
import logging
from pathlib import Path
from dotenv import find_dotenv, load_dotenv
import openpyxl as op
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from src.models.constants import Constants
import re

@click.command()
@click.argument('input_filepath', type=str)
@click.argument('output_filename', type=str)
@click.option('--data_path', default="./data",type=click.Path(exists=True))
@click.option('--preprocess', default=True, type=bool)
@click.option('--cleanup', default=False, type=bool)
@click.option('--averaging', default=False, type= bool)
def main(input_filepath, output_filename,data_path,preprocess,cleanup,averaging):
    """ Runs data processing scripts to turn raw data from (../raw) into
        cleaned data ready to be analyzed (saved in ../processed).
        Options: 
                output_filename = string name of output files (must be a .csv) eg 002-to-011-skip-Mn-Zn
                data_path = string file directory where data is accessible. Expects
                    raw data to be under data_path / raw
                    processed data to be stored under data_path / processed
                preprocess = bool that treats the input filepath as completely raw data to process
                cleanup = bool that applies cleanup to partially processed data
                averaging = bool that applies averaging to cleaned up data
    """

    if make:
        make_file_path = data_path + "/processed/" + "1-preprocessed_"+ output_filename + ".csv"
        raw_data_path = data_path + "/raw"
        df_output=make(input_filepath,raw_data_path=raw_data_path)
        df_output.to_csv(make_file_path)
        print("\nSuccessfully wrote matrix to {}\n".format(make_file_path))
        print("Preview of matrix:\n\n\n{}".format(df_output.head()))
        input_filepath = make_file_path #Change the filepath to use the new one.

    if cleanup:
        cleanup_file_path = data_path + "/processed/" + "2-cleanedup_"+ output_filename + ".csv"
        df_cleanup = post_process(input_filepath,cleanup_file_path,cleanup=True,averaging =False)
        print(f'After stripping, {len(df_cleanup)} catalyst samples in matrix.')
        print(f'Writing to {cleanup_file_path}')
        df_cleanup.to_csv(cleanup_file_path)
        input_filepath = cleanup_file_path #Change the filepath to use the new one.

    if averaging:
        averaging_file_path = data_path + "/processed/" + "3-averaged_"+ output_filename + ".csv"
        df_averaging = post_process(input_filepath,averaging_file_path,cleanup=False,averaging =True)
        print(f'After averaging, {len(df_averaging)} catalyst samples in matrix.')
        print(f'Writing to {averaging_file_path}')
        df_averaging.to_csv(averaging_file_path)

    logger = logging.getLogger(__name__)
    logger.info('making final data set from raw data')

def calc_reactor_metrics(cat,timepts=(0,240)):
    """This function computes catalyst performance metrics by performing a linear regression on yield vs time data. See implementation details
    for a detailed explanation of these choices.
    Input: cat --> Dictionary extracted from raw data.
           timepts: --> Tuple of timepoints to calculate non-lifetime yield"""
    cols = ["Time Since Start","Conversion","Selectivity","Carbon Balance"] #construct matrix of time-on-stream values
    df = pd.DataFrame({i:cat[i] for i in cols})
    df = df[(df["Carbon Balance"] < 1.05) & (df["Carbon Balance"] > 0.95)] #restrict to trustworthy datapoints
    df["Yield"] = df["Conversion"] * df["Selectivity"]
    df["Ln Yield"] = np.log(df["Yield"])

    #perform regression
    reg = LinearRegression().fit(df["Time Since Start"].to_numpy().reshape(-1,1),df["Ln Yield"])

    Y0 = np.exp(reg.intercept_)
    k_d = reg.coef_[0]
    Y_1c = Y0/-k_d

    t0 = timepts[0]
    tf = timepts[1]
    Y_pc = -Y_1c * (np.exp(k_d*tf)-np.exp(k_d*t0))
    sqrtY0Y_pc = np.sqrt(Y0*Y_pc)

    return (-k_d,Y0,Y_1c,Y_pc,sqrtY0Y_pc)

def metal_to_pt_ratio(metal,name):
    """This function returns the molar ratio of a given metal to Pt in a catalyst based on its name.
    Input: metal --> String of a metal
    Input: name --> String of a catalyst name
    """
    metal=metal.lower()
    name=name.lower()
    if len(name.split(metal))==1:
        return 0
    elif len(name.split(metal)[1])==1: #if metal of choice has an integer ratio and is last element
        return int(name.split(metal)[1][0])
    elif '.' == name.split(metal)[1][1]: #if metal of choice has a decimal ratio
        return re.findall("\d+\.\d+", name.split(metal)[1])[0]
    elif len(name.split(metal))==2: #if metal does not have a decimal ratio
        return int(name.split(metal)[1][0])
    else:
        raise ValueError(f'{name} has too many instances of {metal}')


def make_cat_dict(input_filepath,raw_data_path="../../data/raw"):

    raw_data_files = os.listdir(raw_data_path)  # list of all files in the raw data directory

    catalysts = []
    
    with open(raw_data_path+"/"+"Runs_To_Analyze/"+input_filepath,'r') as f:
        for rxn in f.readlines()[1:]: # skip the header
            rxn=rxn.strip()
            try:
                rxn_dirname = [x for x in raw_data_files if rxn in x][0] # select the directory for the reaction
                rxn_fname = [x for x in os.listdir(raw_data_path+"/"+rxn_dirname) if "Analysis6Flow" in x][0]  #select the 6-flow analysis file
                workbook = op.load_workbook(filename=(raw_data_path+"/"+rxn_dirname+"/"+rxn_fname),data_only=True)
                ws = workbook["Summary"]
                print(f'Loading in reaction {rxn}')
            except IndexError as e:
                print("No data for reaction ",rxn)
                continue

            for tube in range(1,7):
                tube_mult_header = 3*(tube-1)
                tube_mult_body = 6*(tube-1)
                n_data_points = 100
                col_letter = op.utils.get_column_letter(tube_mult_header+3)
                if ws[f'{col_letter}2'].value == None: # if the tube is empty, skip it
                    continue
                else:
                    cat_dict = {"Date": ws["A2"].value,
                    "Reaction" : ws["B2"].value,
                    "Tube" : tube,
                    "Catalyst" : ws[f'{op.utils.get_column_letter(tube_mult_header+3)}2'].value,
                    "Mass Catalyst" : ws[f'{op.utils.get_column_letter(tube_mult_header+4)}2'].value,
                    "Mass Diluent" : ws[f'{op.utils.get_column_letter(tube_mult_header+5)}2'].value,
                    "Time":[ws[f'{op.utils.get_column_letter(tube_mult_body+2)}{i}'].value for i in range(10,10+n_data_points)],
                    "Time Since Start":[ws[f'{op.utils.get_column_letter(tube_mult_body+3)}{i}'].value for i in range(10,10+n_data_points)],
                    "Conversion":[ws[f'{op.utils.get_column_letter(tube_mult_body+4)}{i}'].value for i in range(10,10+n_data_points)] ,
                    "Selectivity":[ws[f'{op.utils.get_column_letter(tube_mult_body+5)}{i}'].value for i in range(10,10+n_data_points)],
                    "Carbon Balance":[ws[f'{op.utils.get_column_letter(tube_mult_body+6)}{i}'].value for i in range(10,10+n_data_points)]        
                    }
                    for i in range(n_data_points): # remove the #N/A values from the end of the data
                        if cat_dict["Conversion"][i] == '#DIV/0!':
                            remove_this_and_beyond = i
                            break
                    for col_name in ["Time","Time Since Start","Conversion","Selectivity","Carbon Balance"]:
                        cat_dict[col_name] = cat_dict[col_name][:remove_this_and_beyond]

                    catalysts.append(cat_dict)
    return catalysts   

def make(input_filepath,raw_data_path="../../data/raw"):
    """ Runs data processing scripts to turn raw data from (../raw) into
        cleaned data ready to be analyzed (saved in ../processed).
    """
    catalysts = make_cat_dict(input_filepath,raw_data_path=raw_data_path)
    cols = ["Catalyst","Reaction","Tube","Date","Mass Catalyst","Mass Diluent"]
    df = pd.DataFrame({i:[catalysts[j][i] for j in range(len(catalysts))] for i in cols})
    performance_metrics = [calc_reactor_metrics(cat) for cat in catalysts]
    df["k_d"] = [i[0] for i in performance_metrics]
    df["Y0"] = [i[1] for i in performance_metrics]
    df["lifetime_yield"] = [i[2] for i in performance_metrics]
    df["Y_pc"] = [i[3] for i in performance_metrics]
    df["sqrtY0Y_pc"] = [i[4] for i in performance_metrics]

    for metal in Constants().ALL_TESTED_METALS:
        df[metal] = [metal_to_pt_ratio(metal,name) for name in df["Catalyst"]]

    return df
    
def post_process(infile,outfile,cleanup=True,averaging =True):
    df = pd.read_csv(infile,index_col=0)
    if cleanup:
        print("Performing data cleanup")
        print(f'Currently {len(df)} catalyst samples in matrix.')

        #Remove catalysts with unknown metals to start
        df = df[df["Mn"] == 0]
        df = df[df["Zn"] == 0]

        #Remove catalysts used for testing of transport limitations / dilution effects / background reactivity
        blacklist = ["SiC (new)",
                     "Pt1Sn4Ca4/Al2O3 QS",
                     "Pt1Sn4Ca4/Al2O3 MB",
                     "Pt1Sn4Ca4/Al2O3 (0.3)",
                     "Pt1Sn4Ca4/Al2O3 (0.9)",
                     "Pt1Sn4Ca4/Al2O3 (1.9)",
                     "Pt1Sn1Ga1Fe1Cu1Ca1/Al2O3",
                     "Pt1Sn4Ga1Fe4Cu4Ca4/Al2O3",
                     "Pt1Fur"
                    ]

        df = df[~df["Catalyst"].isin(blacklist)]

        #Remove specific catalysts that are outliers
        idxs = []
        for i, row in df.iterrows():
            if row["Catalyst"] == "Pt1Ga1/Al2O3 180-425um " and row["Reaction"] == "24-003" and row["Tube"] == 2:
                idxs.append(i)
            elif  row["Catalyst"] == "Pt1Sn4Ca4/Al2O3 (1.4)" and row["Reaction"] == "24-008" and row["Tube"] != 1: #Preserve tube 1 but remove the rest
                idxs.append(i)

        df=df.drop(idxs)
        return df

    if averaging:
        print("Performing averaging")
        print(f'Currently {len(df)} catalyst samples in matrix.')
        metals_to_avg_on = Constants().ALL_TESTED_METALS

        df_avg = df.groupby(metals_to_avg_on)[Constants().METRICS].mean().reset_index()
        sd_pct_lifetime = .176
        sd_pct_Y_pc = .033/2
        sd_pct_sqrtY0Y_pc = .033/2 #placeholder
        df_avg["lifetime_yield_sd"] = df_avg["lifetime_yield"]*sd_pct_lifetime #producing an estimated SD for a parameter
        df_avg["Y_pc_sd"] = df_avg["Y_pc"]*sd_pct_Y_pc #producing an estimated SD for a parameter
        df_avg["sqrtY0Y_pc_sd"] = df_avg["sqrtY0Y_pc"]*sd_pct_sqrtY0Y_pc
        return df_avg

    
    logger = logging.getLogger(__name__)
    logger.info('making final data set from raw data')


if __name__ == '__main__':
    log_fmt = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=log_fmt)

    # not used in this stub but often useful for finding various files
    project_dir = Path(__file__).resolve().parents[2]

    # find .env automagically by walking up directories until it's found, then
    # load up the .env entries as environment variables
    load_dotenv(find_dotenv())

    main()
